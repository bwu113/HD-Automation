import xlwings as xw
from xlwings.constants import HAlign
import datetime as dt
import tkinter as tk
from openpyxl import utils as opyxl
from tkinter import simpledialog
from tkinter import filedialog
from tkinter import messagebox

def insertExcel(PackingList):
    exlFilePath = filedialog.askopenfilename(title="Select Excel File To Process")
    exlSheet = simpledialog.askstring(" ", prompt="Enter Sheet Name")
    wb = xw.Book(exlFilePath)
    while exlSheet not in wb.sheet_names:
        messagebox.showinfo("Error", "Sheet doesn't exist, Make sure sheet name is correct")
        exlSheet = simpledialog.askstring(" ", prompt="Enter Sheet Name")
    wks = xw.sheets(exlSheet)
    maxCol = 700
    rowSpaceConst = 52
    currCol = 0
    currRow = 0
    i = 0
    PrintArea = ""
    ROOT = tk.Tk()
    ROOT.withdraw()
    ErrNames = []

    rowRange = wks[2,10]

    print(rowRange.value)
    # Salesman = simpledialog.askstring(title=" ", prompt="Enter Your Name")
    # invoice = int(simpledialog.askstring(title=" ", prompt="Enter Invoice # to start from"))
    # if(invoice):
    #     msgConfirm = messagebox.askyesno("Confirm", "Are you sure " + str(invoice) + " is the correct invoice number?")
    #     while msgConfirm is False: #Confirming invoice number
    #         invoice = simpledialog.askstring(title=" ", prompt="Enter Invoice # to start from")
    #         msgConfirm = messagebox.askyesno("Confirm", "Are you sure " + str(invoice) + " is the correct invoice number?")
    #     startingCol = simpledialog.askstring(title=" ", prompt="Enter starting column letter(s)").upper()
    #     msgConfirm = messagebox.askyesno("Confirm","Is the starting column: " + startingCol + " correct?")
    #     while msgConfirm is False: #Confirming starting Column Letter
    #         startingCol = simpledialog.askstring(title=" ", prompt="Enter starting column letter(s)").upper()
    #         msgConfirm = messagebox.askyesno("Confirm","Is the starting column: " + startingCol + " correct?")
    #     startingRow = simpledialog.askstring(title=" ", prompt="Enter starting row number")
    #     msgConfirm = messagebox.askyesno("Confirm","Is the starting row: " + startingRow + " correct?")
    #     while msgConfirm is False: #Confirming starting Row number
    #         startingRow = simpledialog.askstring(title=" ", prompt="Enter starting row number")
    #         msgConfirm = messagebox.askyesno("Confirm","Is the starting row: " + startingRow + " correct?")
    #     currCol = ConvertColumn(startingCol) + 1
    #     if(currCol <= (maxCol-7)):
    #         if startingRow == 0:
    #             currRow = 7
    #         else:
    #             currRow = int(startingRow) + 6
    #     else:
    #         currCol = 2
    #         if startingRow == 0:
    #             currRow = 7
    #         else:
    #             currRow = int(startingRow) + 6

    #     for item in PackingList.values():
    #         # if("err" in item):
    #         #     ErrNames.append(item["name"])
    #         #     continue
    # #Delivery Information Labels
    #         wks.range(currRow,currCol).value = "SOLD TO :"
    #         wks.range(currRow-1,currCol+1).value = "Home Depot"
    # #Delivery Information
    #         if("storenum" in  item):
    #             wks.range(currRow,currCol+1).value = [[item["name"]],[item["storenum"]],[item["street"]],[item["cityzip"]],[item["phone"]]]
    #         elif("sean" in item["orderby"].lower()):
    #             wks.range(currRow,currCol+1).value = [[item["name"]],[item["street"]],[item["cityzip"]]]
    #         else:
    #             wks.range(currRow,currCol+1).value = [[item["name"]],[item["street"]],[item["cityzip"]],[item["phone"]]]
    # #Invoice / Date /Salesman Labels
    #         if("sean" in item["orderby"].lower()):
    #             wks.range(currRow-1,currCol+4).value = "HD Connect"
    #         wks.range(currRow,currCol+3).value = [["INVOICE #"],["DATE:"],["SALESMAN:"],["TERMS:"],["SHIP VIA:"]]
    # #Invoice / Date / Salesman
    #         wks.range(currRow,currCol+4).value = "HD-" + str(invoice)
    #         wks.range(currRow+1,currCol+4).value = dt.datetime.today().strftime('%m/%d/%Y')
    #         wks.range(currRow+2,currCol+4).value = Salesman
    # #Order Info Headers
    #         wks.range(currRow+7,currCol+2).value = "INVOICE"
    #         wks.range(currRow+8,currCol-1).value = ["QTY","UNIT","Item #","DESCRIPTION","UN. PRICE","AMOUNT"]
    #         untPrice = simpledialog.askstring(title=" ", prompt=item['name'] + "\nEnter Unit Price")
    #         while(untPrice == "" or untPrice is None):
    #             untPrice = simpledialog.askstring(title=" ", prompt=item['name'] + "\nEnter Unit Price")
    #         totalPrice = float(untPrice) * float(item['qty'])
    # #Order Information
    #         #Enters normally for single sku orders with no addons
    #         wks.range(currRow+10,currCol-1).value = [item["qty"],"pcs",item["sku"].rstrip("-"),item["desc"], untPrice, totalPrice]
    #         #Enters Backsplash or Mirror addons
    #         if("bs" in item and "mir" in item):
    #             wks.range(currRow+11,currCol-1).value = [item["qty"],"pcs","BS","Matching Backsplash","Inclusive",""]
    #             wks.range(currRow+12,currCol-1).value = [item["qty"],"pcs","MIR","Matching Mirror","Inclusive",""]
    #         elif("bs" in item or "mir" in item):
    #             if("bs" in item):
    #                 wks.range(currRow+11,currCol-1).value = [item["qty"],"pcs","BS","Matching Backsplash","Inclusive",""]
    #             elif("mir" in item):
    #                 wks.range(currRow+11,currCol-1).value = [item["qty"],"pcs","MIR","Matching Mirror","Inclusive",""]  
    #         if("sku2" in item):
    #             untPrice = simpledialog.askstring(title="Second Unit Price", prompt="Enter Unit Price For Second SKU")
    #             while(untPrice == "" or untPrice is None):
    #                 untPrice = simpledialog.askstring(title="Second Unit Price", prompt="Enter Unit Price For Second SKU")
    #             totalPrice = float(untPrice) * float(item['qty2'])
    #             if(not wks.range(currRow+11,currCol-1).value): #Checks if row below 1st sku is empty or not proceeds if its empty.
    #                 wks.range(currRow+11,currCol-1).value = [item["qty2"],"pcs",item["sku2"].rstrip("-"),item["desc2"], untPrice, totalPrice]

    # #Tracking Info
    #         #Left Labels
    #         wks.range(currRow+18,currCol).value = "Trucking(LTL) company"
    #         wks.range(currRow+19,currCol+1).value = "PRO #"
    #         wks.range(currRow+20,currCol).value = "Shipping Service"
    #         #Right Half
    #         wks.range(currRow+20,currCol+2).value = "LTL"
    # #Customer Order and PO
    #         #Left Labels
    #         wks.range(currRow+25,currCol+1).value = [["Customer Order #"],["PO #"]]
    #         #Right half        
    #         wks.range(currRow+25,currCol+2).value = item["co"]
    #         if item["po"][0] == "0":
    #             wks.range(currRow+26,currCol+2).value = "'" + item["po"]
    #         else:
    #             wks.range(currRow+26,currCol+2).value = item["po"]
    # #Sub Total and Totals
    #         #Left Labels
    #         wks.range(currRow+35,currCol+3).value = [["Sub. Total"],["Sales Tax"],["Total"]]
    #         #Right Half
    #         if("sku2" in item):
    #             subTotal = float(wks.range(currRow+10,currCol+4).value) + float(wks.range(currRow+11,currCol+4).value)
    #         else:
    #             subTotal = wks.range(currRow+10,currCol+4).value
    #         wks.range(currRow+35,currCol+4).value = subTotal
    #         wks.range(currRow+37,currCol+4).value = subTotal
    #     #Set Print Area
    #         if(i == len(PackingList)-1):
    #             PrintArea += opyxl.cell.get_column_letter(currCol-1) + str(currRow-6) + ":" + opyxl.cell.get_column_letter(currCol+4) + str(currRow+37)
    #         else:
    #             PrintArea += opyxl.cell.get_column_letter(currCol-1) + str(currRow-6) + ":" + opyxl.cell.get_column_letter(currCol+4) + str(currRow+37) + ","
    #     # Set next order's col
    #         currCol += 7
    #         if(currCol > maxCol): #If col would exceed the max column length then reset to 2 and set new starting row
    #             currCol = 2
    #             currRow += rowSpaceConst
    #         #Increment Invoice Number    
    #         invoice += 1
    #         #Increment Counter
    #         i += 1
    # else:
    #     messagebox.showinfo("Canceled","Canceled")
    # wks.page_setup.print_area = PrintArea
    #messagebox.showinfo("Skipped Pages", ErrNames)
def ConvertColumn(column):
    result = 0; 
    for B in range(len(column)): 
        result *= 26; 
        result += ord(column[B]) - ord('A') + 1; 
    return result; 
